"""Convert backtest CSVs to a formatted Excel workbook.

Usage:
  python scripts/to_excel.py --date 2026-05-29 --symbol NIFTY
  python scripts/to_excel.py --date 2026-05-29 --symbol NIFTY --out-dir scripts/output
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import date, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter

import pytz
IST = pytz.timezone("Asia/Kolkata")

# ── Colour palette ─────────────────────────────────────────────────────────────
C_HEADER_BG   = "1A1A2E"   # dark navy
C_HEADER_FG   = "E0E0E0"   # light grey text
C_BUY_BG      = "D6F5D6"   # light green
C_SELL_BG     = "FFD6D6"   # light red
C_EXIT_B_BG   = "EAF7EA"   # very light green
C_EXIT_S_BG   = "FFF0F0"   # very light red
C_SIGNAL_FG   = "1A1A1A"   # dark text on signal rows
C_UP_FG       = "1B5E20"   # dark green for UP direction
C_DOWN_FG     = "B71C1C"   # dark red for DOWN direction
C_ALT_BG      = "F8F8F8"   # alternating row
C_WHITE_BG    = "FFFFFF"
C_WIN_BG      = "C8E6C9"   # green for profitable trade
C_LOSS_BG     = "FFCDD2"   # red for losing trade
C_SECTION_BG  = "37474F"   # section header
C_SECTION_FG  = "FFFFFF"
C_ACCENT_BG   = "E3F2FD"   # light blue accent for summary

thin = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")


def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


# ── Sheet 1: Bars ─────────────────────────────────────────────────────────────

BAR_COLUMNS = [
    ("Time",      "timestamp",    12, "center"),
    ("Open",      "open",         10, "right"),
    ("High",      "high",         10, "right"),
    ("Low",       "low",          10, "right"),
    ("Close",     "close",        10, "right"),
    ("SuperTrend","st",           10, "right"),
    ("ST Dir",    "st_dir",        7, "center"),
    ("EMA 21",    "ema21",        10, "right"),
    ("PSAR",      "psar",         10, "right"),
    ("ATR",       "atr",           8, "right"),
    ("RSI",       "rsi",           8, "right"),
    ("Prev Mid",  "prev_mid",      9, "right"),
    ("Signal",    "signal",        9, "center"),
    ("Position",  "position",      9, "center"),
    ("C1",        "c1",            5, "center"),
    ("C2",        "c2",            5, "center"),
    ("C3",        "c3",            5, "center"),
    ("C4",        "c4",            5, "center"),
    ("C5 RSI",    "c5",            6, "center"),
]

def write_bars_sheet(ws, df_bars: pd.DataFrame, symbol: str, target_date: str):
    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:S1")
    title_cell = ws["A1"]
    title_cell.value = f"{symbol}  5-Min Bars  —  {target_date}  |  Strategy: SuperTrend + PSAR + EMA21 + RSI"
    title_cell.font  = font(bold=True, color=C_HEADER_FG, size=13)
    title_cell.fill  = fill(C_HEADER_BG)
    title_cell.alignment = align("center")
    ws.row_dimensions[1].height = 24

    # ── Legend ────────────────────────────────────────────────────────────────
    ws.merge_cells("A2:S2")
    legend = ws["A2"]
    legend.value = (
        "Green = BUY signal   |   Red = SELL signal   |   "
        "C1-C5 = entry conditions (True/False)   |   "
        "ST Dir UP = uptrend   DOWN = downtrend"
    )
    legend.font      = font(italic=True, color="555555", size=9)
    legend.fill      = fill("F0F0F0")
    legend.alignment = align("center")
    ws.row_dimensions[2].height = 16

    # ── Condition legend ──────────────────────────────────────────────────────
    ws.merge_cells("A3:S3")
    cond_leg = ws["A3"]
    cond_leg.value = (
        "C1: price vs ST+0.2*ATR   |   "
        "C2: not overextended (within 1 ATR of ST)   |   "
        "C3: price vs PSAR   |   "
        "C4: price vs EMA21   |   "
        "C5: RSI > 50 (BUY) / RSI < 50 (SELL)"
    )
    cond_leg.font      = font(italic=True, color="555555", size=9)
    cond_leg.fill      = fill("F0F0F0")
    cond_leg.alignment = align("center")
    ws.row_dimensions[3].height = 16

    # ── Column headers ─────────────────────────────────────────────────────────
    header_row = 4
    for col_idx, (label, _, width, halign) in enumerate(BAR_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font      = font(bold=True, color=C_HEADER_FG, size=10)
        cell.fill      = fill(C_SECTION_BG)
        cell.alignment = align(halign)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, (_, row) in enumerate(df_bars.iterrows(), start=header_row + 1):
        signal   = str(row.get("signal", "-") or "-").strip()
        position = str(row.get("position", "FLAT") or "FLAT").strip()
        st_dir   = str(row.get("st_dir", "") or "").strip()

        # Row background
        if signal == "BUY":
            row_bg = C_BUY_BG
        elif signal == "SELL":
            row_bg = C_SELL_BG
        elif signal == "BUY_EXIT":
            row_bg = C_EXIT_B_BG
        elif signal == "SELL_EXIT":
            row_bg = C_EXIT_S_BG
        elif row_idx % 2 == 0:
            row_bg = C_ALT_BG
        else:
            row_bg = C_WHITE_BG

        ws.row_dimensions[row_idx].height = 16

        for col_idx, (_, field, _, halign) in enumerate(BAR_COLUMNS, start=1):
            # Resolve value
            raw = row.get(field, "")
            if field == "timestamp":
                # Show only HH:MM
                val = str(raw)[11:16] if raw else ""
            elif field in ("c1", "c2", "c3", "c4", "c5"):
                # Try all possible column name variants
                for key in (field, f"c1_above_st", f"c1_below_st",
                             f"c2_not_over", f"c3_psar", f"c4_ema", f"c5_rsi"):
                    if key in row.index:
                        raw = row[key]
                        break
                # Re-resolve c1–c5 by position
                cond_keys = [k for k in row.index if k.startswith("c") and k[1:2].isdigit()]
                cond_keys.sort()
                c_pos = col_idx - 15   # columns 15-19 = c1-c5
                if 0 <= c_pos < len(cond_keys):
                    raw = row[cond_keys[c_pos]]
                val = "Y" if str(raw).lower() == "true" else ("N" if str(raw).lower() == "false" else "")
            elif field == "signal":
                val = signal if signal not in ("-", "WARMUP", "NAN") else ""
            else:
                val = raw

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = fill(row_bg)
            cell.border    = BORDER
            cell.alignment = align(halign)

            # Special colours
            if field == "st_dir":
                if st_dir == "UP":
                    cell.font = font(bold=True, color=C_UP_FG)
                elif st_dir == "DOWN":
                    cell.font = font(bold=True, color=C_DOWN_FG)
                else:
                    cell.font = font()
            elif field == "signal" and val:
                cell.font = font(bold=True, color="1A1A1A", size=10)
            elif field in ("c1", "c2", "c3", "c4", "c5"):
                if val == "Y":
                    cell.font = font(bold=True, color="1B5E20")
                elif val == "N":
                    cell.font = font(color="B71C1C")
                else:
                    cell.font = font()
            elif field == "rsi":
                v = float(val) if str(val).replace(".", "").replace("-", "").isdigit() else None
                if v is not None:
                    if v > 70:
                        cell.font = font(bold=True, color="1B5E20")
                    elif v < 30:
                        cell.font = font(bold=True, color="B71C1C")
                    else:
                        cell.font = font()
                else:
                    cell.font = font()
            else:
                cell.font = font()

    # ── Freeze top rows ───────────────────────────────────────────────────────
    ws.freeze_panes = "A5"


# ── Sheet 2: Trades ───────────────────────────────────────────────────────────

TRADE_COLUMNS = [
    ("Side",         "side",         8,  "center"),
    ("Entry Time",   "entry_time",   22, "center"),
    ("Entry Price",  "entry_price",  12, "right"),
    ("Exit Time",    "exit_time",    22, "center"),
    ("Exit Price",   "exit_price",   12, "right"),
    ("P&L (pts)",    "pnl_pts",      11, "right"),
    ("P&L (%)",      "pnl_pct",      10, "right"),
    ("Result",       "result",        9, "center"),
    ("Exit Reason",  "exit_reason",  40, "left"),
]

def write_trades_sheet(ws, df_trades: pd.DataFrame, symbol: str, target_date: str):
    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    tc = ws["A1"]
    tc.value = f"{symbol}  Trade Log  —  {target_date}"
    tc.font  = font(bold=True, color=C_HEADER_FG, size=13)
    tc.fill  = fill(C_HEADER_BG)
    tc.alignment = align("center")
    ws.row_dimensions[1].height = 24

    if df_trades.empty:
        ws.merge_cells("A2:I2")
        empty_cell = ws["A2"]
        empty_cell.value = "No completed trades on this date."
        empty_cell.font  = font(italic=True, color="888888")
        empty_cell.alignment = align("center")
        return

    # ── Summary strip ─────────────────────────────────────────────────────────
    total    = len(df_trades)
    wins     = len(df_trades[df_trades["result"] == "PROFIT"])
    losses   = total - wins
    total_pnl = df_trades["pnl_pts"].sum()
    win_rate  = wins / total * 100 if total > 0 else 0

    summary_items = [
        ("Total trades", total),
        ("Wins",         wins),
        ("Losses",       losses),
        ("Win rate",     f"{win_rate:.1f}%"),
        ("Net P&L (pts)",f"{total_pnl:+.2f}"),
    ]
    for col_idx, (label, val) in enumerate(summary_items, start=1):
        lc = ws.cell(row=2, column=col_idx * 2 - 1, value=label)
        vc = ws.cell(row=2, column=col_idx * 2,     value=val)
        lc.font      = font(bold=True, color="FFFFFF", size=9)
        lc.fill      = fill(C_SECTION_BG)
        lc.alignment = align("center")
        vc.font      = font(bold=True, color="FFFFFF", size=11)
        vc.fill      = fill(C_SECTION_BG)
        vc.alignment = align("center")
        ws.row_dimensions[2].height = 22

    # ── Headers ───────────────────────────────────────────────────────────────
    header_row = 3
    for col_idx, (label, _, width, halign) in enumerate(TRADE_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font      = font(bold=True, color=C_HEADER_FG)
        cell.fill      = fill(C_SECTION_BG)
        cell.alignment = align(halign)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, (_, row) in enumerate(df_trades.iterrows(), start=header_row + 1):
        result = str(row.get("result", "")).strip()
        row_bg = C_WIN_BG if result == "PROFIT" else C_LOSS_BG

        ws.row_dimensions[row_idx].height = 18

        for col_idx, (_, field, _, halign) in enumerate(TRADE_COLUMNS, start=1):
            raw = row.get(field, "")
            if field in ("entry_time", "exit_time"):
                val = str(raw)[:19] if raw else ""
            elif field == "pnl_pts":
                val = float(raw) if raw != "" else ""
            elif field == "pnl_pct":
                val = float(raw) if raw != "" else ""
            else:
                val = raw

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = fill(row_bg)
            cell.border    = BORDER
            cell.alignment = align(halign)

            if field == "result":
                cell.font = font(bold=True,
                                 color=C_UP_FG if result == "PROFIT" else C_DOWN_FG)
            elif field == "pnl_pts":
                cell.font = font(bold=True,
                                 color=C_UP_FG if (val or 0) > 0 else C_DOWN_FG)
                if isinstance(val, float):
                    cell.number_format = '+0.00;-0.00'
            elif field == "pnl_pct":
                if isinstance(val, float):
                    cell.number_format = '+0.000%;-0.000%'
                    cell.value = val / 100  # Excel % format
                cell.font = font(color=C_UP_FG if (val or 0) > 0 else C_DOWN_FG)
            elif field == "side":
                cell.font = font(bold=True,
                                 color=C_UP_FG if str(raw) == "LONG" else C_DOWN_FG)
            else:
                cell.font = font()

    ws.freeze_panes = "A4"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(description="Backtest CSV to Excel")
    p.add_argument("--symbol",  default="NIFTY")
    p.add_argument("--date",    default="")
    p.add_argument("--out-dir", default="scripts/output")
    args = p.parse_args()

    if args.date:
        target_date = args.date
    else:
        from datetime import datetime
        import pytz
        ist = pytz.timezone("Asia/Kolkata")
        d = datetime.now(ist).date() - timedelta(days=1)
        while d.weekday() >= 5:
            d -= timedelta(days=1)
        target_date = d.isoformat()

    sym    = args.symbol.upper()
    sym_lo = sym.lower()

    bars_path   = os.path.join(args.out_dir, f"{sym_lo}_bars_{target_date}.csv")
    trades_path = os.path.join(args.out_dir, f"{sym_lo}_trades_{target_date}.csv")
    out_path    = os.path.join(args.out_dir, f"{sym_lo}_backtest_{target_date}.xlsx")

    # Load CSVs
    if not os.path.exists(bars_path):
        print(f"ERROR: {bars_path} not found. Run backtest.py first.")
        sys.exit(1)

    df_bars = pd.read_csv(bars_path)
    df_trades = pd.read_csv(trades_path) if os.path.exists(trades_path) else pd.DataFrame()

    print(f"Building Excel: {sym} {target_date}")
    print(f"  Bars:   {len(df_bars)} rows")
    print(f"  Trades: {len(df_trades)} rows")

    wb = Workbook()

    # Sheet 1: Bars
    ws_bars = wb.active
    ws_bars.title = f"{sym} Bars {target_date}"
    ws_bars.sheet_view.showGridLines = False
    write_bars_sheet(ws_bars, df_bars, sym, target_date)

    # Sheet 2: Trades
    ws_trades = wb.create_sheet(title=f"{sym} Trades {target_date}")
    ws_trades.sheet_view.showGridLines = False
    write_trades_sheet(ws_trades, df_trades, sym, target_date)

    wb.save(out_path)
    print(f"  saved: {out_path}")
    print("Done.")


if __name__ == "__main__":
    main()
